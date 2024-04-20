import time
import os
import http
import requests
from flask import Flask, request
from werkzeug.wrappers import Response
import telebot
from telebot import types, apihelper
import requests
import base64
from google.cloud import storage
import openpyxl
from openpyxl import load_workbook
import vertexai
from vertexai.generative_models import GenerativeModel, Part, Image
# from PIL import Image
import http.client
import typing
import urllib.request
import json
import tempfile

app = Flask(__name__)
OPEN_API_KEY = os.getenv("OPENAI_API_KEY")
BOT_TOKEN = os.environ.get('BOT_TOKEN')

bot = telebot.TeleBot(BOT_TOKEN)
# Initialize GCS client
client = storage.Client()
selected_file = None
destination_bucket_name = "cloud_computing_excel_buckets"
destination_file_name = "Budget.xlsx"

@app.post("/")
def index() -> Response:
    # dispatcher.process_update(
    #     Update.de_json(request.get_json(force=True), bot))
    if request.method == 'POST':
        update = telebot.types.Update.de_json(request.get_json(force=True))
        bot.process_new_updates([update])
        return f"Status Code: 200"
    return "Not ok"

@bot.message_handler(commands=['start', 'hello', 'spreadsheet'])
def start(message):
    bot.reply_to(message, "Please wait as I create the markup keyboard...")
    show_main_markup(message.chat.id)

def show_main_markup(chat_id):
    markup = send_welcome()
    bot.send_message(chat_id, "Choose an option:", reply_markup=markup)

def send_welcome():
    markup = types.InlineKeyboardMarkup()
    spreadsheet_button = types.InlineKeyboardButton("Spreadsheets", callback_data='spreadsheets')
    expense_button = types.InlineKeyboardButton('Track Expenses', callback_data='report_expense')
    help_button = types.InlineKeyboardButton('Help', callback_data='help')
    markup.add(spreadsheet_button, expense_button, help_button)
    return markup

def create_options_markup():
    markup = types.InlineKeyboardMarkup()
    create_new_button = types.InlineKeyboardButton("Create a New Spreadsheet", callback_data='create_new')
    fetch_existing_button = types.InlineKeyboardButton("Fetch Existing Spreadsheets", callback_data='fetch_existing')
    back_button = types.InlineKeyboardButton("Back", callback_data='back_to_main')
    markup.add(create_new_button, fetch_existing_button)
    markup.row(back_button)
    return markup

# Handle all other messages
@bot.message_handler(func=lambda message: True, content_types=['text'])
def echo_message(message):
    bot.reply_to(message, "Hi, I'm a chatbot powered by Gemini AI to help track your expenses. Please start by using /spreadsheet to select a spreadsheet to track your expenses in!")

@bot.message_handler(commands=['expense'])
def start_expense(message):
    bot.reply_to(message, "Please send me the receipt image. Make sure you already selected a spreadsheet or created one before this!!")

@bot.callback_query_handler(func=lambda call: True)
def handle_query(call):
    if call.data == 'spreadsheets':
        options_markup = create_options_markup()
        bot.send_message(call.message.chat.id, "Choose an option:", reply_markup=options_markup)

    elif call.data == 'create_new':
        bot.send_message(call.message.chat.id, "Creating a new spreadsheet...")
        # Add your logic for creating a new spreadsheet here
        # Set the name of the GCS bucket and the file to copy
        source_bucket_name = "cloud_computing_excel_buckets"
        source_file_name = "Budget Template.xlsx"
        destination_bucket_name = "cloud_computing_excel_buckets"
        destination_file_name = "Budget.xlsx"

        # Get source bucket and blob
        source_bucket = client.bucket(source_bucket_name)
        source_blob = source_bucket.blob(source_file_name)

        # Get destination bucket
        destination_bucket = client.bucket(destination_bucket_name)

        # Copy the file to the destination bucket
        source_bucket.copy_blob(source_blob, destination_bucket, destination_file_name)

        # Send a response to the user
        bot.send_message(call.message.chat.id, "Excel sheet created successfully! Please send me the receipt image.")

    elif call.data == 'fetch_existing':
        bot.send_message(call.message.chat.id, "Fetching existing spreadsheets...")
        destination_bucket_name = "cloud_computing_excel_buckets"
        destination_file_name = "Budget.xlsx"
        bucket = client.bucket(destination_bucket_name)
        # List all spreadsheets in the bucket
        blobs = bucket.list_blobs()
        markup = types.InlineKeyboardMarkup()
        
        file_count = 0
        # Create buttons for each spreadsheet
        for blob in blobs:
            if blob.name.endswith('.xlsx'):  # Filter for Excel files first
                file_count += 1
                if file_count > 1:  # Start adding buttons from the second file
                    button = types.InlineKeyboardButton(blob.name, callback_data='select_' + blob.name)
                    markup.add(button)
        
        # Send a message with the inline keyboard
        bot.send_message(call.message.chat.id, "Select a spreadsheet:", reply_markup=markup)

    elif call.data.startswith('select_'):
        selected_file = call.data[len('select_'):]
        # Add logic for what happens when a file is selected
        bot.send_message(call.message.chat.id, f"You selected: {selected_file}. Please send a receipt image to track your expenses in this spreadsheet.")

    # elif call.data == 'report_expense':
    #     # Here you can prompt the user to send a receipt or enter expense details
    #     bot.send_message(call.message.chat.id, "Please send me a receipt image or enter your expense details.")
    elif call.data == 'help':
        # Send a help message or display more options
        bot.send_message(call.message.chat.id, "Here's how you can use me: /spreadsheet to select a spreadsheet for expense tracking. /help to get help.")
    elif call.data == 'back_to_main':
        show_main_markup(call.message.chat.id)


@bot.message_handler(content_types=['photo'])
def handle_image(message):
    if selected_file is None:
        bot.send_message(message.chat.id, "Please select a spreadsheet first using /spreadsheet.")
    else:
        bot.send_message(message.chat.id, "This will take me 1 minute as I process your receipt. Go grab some coffee in the meantime")
        file_info = bot.get_file(message.photo[-1].file_id)
        receipt = load_image_from_url('https://api.telegram.org/file/bot{0}/{1}'.format(BOT_TOKEN, file_info.file_path))
        expense_info = generate_text(message.chat.id, os.environ.get('GCP_PROJECT_ID'), "asia-southeast1", receipt)
        # bot.send_message(message.chat.id, "Thank you for waiting! Here is the expense information")
        if expense_info is not None:
            bot.reply_to(message, f"{expense_info}")
            # Parse the JSON string
            expense_data_json = json.loads(expense_info)
            bot.send_message(message.chat.id, f"Lemme add this to the spreadsheet: {selected_file}")
            # Download, modify, and upload the file
            local_path = download_file_from_gcs(message.chat.id, selected_file)
            # Call function to modify the spreadsheet
            bot.send_message(message.chat.id, f"Adding expense data to {selected_file}...")
            add_expense_to_sheet(local_path, "Income and Expenses", expense_data_json)
            bot.send_message(message.chat.id, "Expense data added successfully! Uploading the updated spreadsheet...")
            upload_file_to_gcs(message.chat.id, local_path, selected_file)

            # Send file to the user
            try:
                with open(local_path, 'rb') as file:
                    bot.send_document(message.chat.id, file, caption="Here's your updated spreadsheet.")
            except apihelper.ApiException as e:
                print(f"An error occurred while sending the document: {e}")
            except FileNotFoundError:
                print(f"The file at {local_path} was not found.")
            except Exception as e:
                print(f"An unexpected error occurred: {e}")

            # Clean up the local file
            os.remove(local_path)
        else:
            bot.reply_to(message, "Sorry, I was not able to process the image. Please try again.")
  

def download_file_from_gcs(chat_id, blob_name):
    """
    Download a file from Google Cloud Storage to a temporary file and handle errors.

    Args:
        chat_id (int): Telegram chat ID to send messages to the user.
        blob_name (str): The name of the file in the GCS bucket to be downloaded.

    Returns:
        str: The path to the downloaded temporary file.
    """
    try:
        # Access the specific GCS bucket
        excel_bucket = client.bucket(destination_bucket_name)
        # Access the blob within the GCS bucket
        blob = excel_bucket.blob(blob_name)
        # bot.send_message(chat_id, f"Downloading {blob.name} from GCS...")
        # Get the directory where the script is running
        current_dir = os.getcwd() #os.path.dirname(os.path.abspath(__file__))
        # Create a full path to save the file
        save_path = os.path.join(current_dir, destination_file_name)
        
        # Send a message to the user
        blob.download_to_filename(save_path)
        # blob.download_to_filename(temp_local_filename)
        bot.send_message(chat_id, f"I downloaded {blob_name} to {save_path}.")
        # Close the file descriptor immediately after creating the temp file to avoid leaks
        return save_path
    except Exception as e:
        # Inform the user of the failure via Telegram
        bot.send_message(chat_id, f"Failed to download the file: {e}")
        # Optionally re-raise the exception if further handling is required
        raise

def upload_file_to_gcs(chat_id, local_file_path, blob_name):
    """Upload a file to Google Cloud Storage."""
    excel_bucket = client.bucket(destination_bucket_name)
    blob = excel_bucket.blob(blob_name)
    blob.upload_from_filename(local_file_path)
    bot.send_message(chat_id, f"Uploaded {blob_name} to GCS successfully!")

def add_expense_to_sheet(file_name, sheet_name, expense_data):
    # Load the workbook and sheet
    wb = load_workbook(filename=file_name)
    ws = wb[sheet_name]

    # Find the first empty row in the range B:E starting from row 4
    start_row = 4
    for item in expense_data["Items"]:
        empty_row_found = False
        
        # Iterate over the rows starting from the last checked row
        for row in ws.iter_rows(min_row=start_row, min_col=2, max_col=5):
            # Check if the entire row is empty
            if all(cell.value is None for cell in row):
                empty_row_found = True
                print(f"Appending item: {item}")
                
                # Fill the empty row with item data
                for col, key in enumerate(item.keys(), start=2):
                    row[col-2].value = item[key]
                
                print(f"Item keys appended to row {row[0].row}.")
                # Move to the next row after filling the current empty row
                start_row = row[0].row + 1
                break  # Break out of the inner loop after filling one empty row
        
        # If no empty row was found for the current item, break out of the loop
        if not empty_row_found:
            print("No more empty rows found.")
            break

    # Save the workbook
    wb.save(filename=file_name)
    print(f"Data added to {file_name} in sheet {sheet_name}.")


# create helper function
def load_image_from_url(image_url: str) -> Image:
    with urllib.request.urlopen(image_url) as response:
        response = typing.cast(http.client.HTTPResponse, response)
        image_bytes = response.read()
    return Image.from_bytes(image_bytes)

def generate_text(chat_id, project_id, location, image):
    # Initialize Vertex AI
    vertexai.init(project=project_id, location=location)

    # Load the model
    model = GenerativeModel(model_name= "gemini-1.5-pro-preview-0409") #"gemini-pro-vision")
    bot.send_message(chat_id, "Performing some OCR magic now")
    # Load example image
    # image_content = Part.from_uri(image_url, "image/jpeg")
    categories = [
        "Income",
        "Housing",
        "Utilities",
        "Groceries",
        "Insurance",
        "Phone",
        "Credit Cards",
        "School",
        "Savings",
        "Entertainment",
        "Food",
        "Others"
    ]
    instructions = f"What are the individiual items, and final total amount? Please return me a dictionary of an array containg dictionaries in JSON format. The key of the main dictionary is 'Items' and I want to have the following keys in the child dictionaries in the array: Category, Description, Amount, Notes. Choose the categories from one of the item in the array: {categories}"
    # Query the model
    response = model.generate_content([image, instructions], stream = False)
    # bot.send_message(chat_id, f"Response Attributes: {response.__attributes__}")
    # print(response.text)
    # Extract the generated text
    generated_text = response.text
    cleaned_start = generated_text.replace("```json", "")  # Remove starting ```json
    cleaned_end = cleaned_start.replace("```", "")  # Remove ending ```
    cleaned_response = cleaned_end.strip()
    return cleaned_response

# def upload_image(chat_id, file, image_path):
#     """Uploads an image to a GCS bucket and returns the gs url.

#     Args:
#         bucket_name: The name of the bucket to upload the image to.
#         image_path: The local path of the image file.

#     Returns:
#         The gs url of the uploaded image.
#     """
#     # base64_encoded_image_str = base64_encode(image)
#     bucket_name = "cz4052_image_bucket"
#     storage_client = storage.Client()
#     image_bucket = storage_client.bucket(bucket_name)

#     # Extract filename from the image path
#     image_name = image_path.split("/")[-1]
#     bot.send_message(chat_id, "File Name: {0}".format(image_name))

#     # Create a blob and upload the image
#     blob = image_bucket.blob(image_name)
#     bot.send_message(chat_id, "Uploading image now!")
#     # blob.upload_from_string(image.content, content_type='image/jpeg')
#     # image = Image.open(file.content)
#     # bs = io.BytesIO()
#     # image.save(bs, "jpeg")
#     # blob.upload_from_file(image)
#     blob.upload_from_string(file.content, content_type='image/jpeg')

#     bot.send_message(chat_id, "Image uploaded successfully!")

#     return f"gs://{bucket_name}/{image_name}"

# PROJECT_ID = os.environ.get('GCP_PROJECT_ID')
# generate_text(PROJECT_ID, "us-central1")
